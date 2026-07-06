import csv
import io
import subprocess
import sys
import sqlite3
import ast
import re
import requests
import json

from openpyxl import load_workbook
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Avg, Count, Sum, Q, Prefetch
from django.contrib.auth.models import User
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from django.db import IntegrityError
from .models import Profile, Exam, Result, Question, Choice, Subject, StudentAnswer, TestCase
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes as drf_permission_classes
from rest_framework.response import Response
from .serializers import ExamSerializer, QuestionSerializer, SubjectSerializer, ChoiceSerializer, ResultSerializer, StudentAnswerSerializer, TestCaseSerializer
from rest_framework import viewsets
from rest_framework.views import APIView
from .permissions import IsTeacher, IsTeacherOrReadOnly

def home(request):
    return render(request, 'exams/home.html')


@login_required
def dashboard(request):
    profile = Profile.objects.filter(user=request.user).first()
    is_preview = request.GET.get('preview') == '1'

    if profile and profile.role == 'teacher' and not is_preview:
        return redirect('teacher_dashboard')
    query = request.GET.get('q', '').strip()
    now = timezone.now()

    exams_queryset = Exam.objects.filter(is_active=True).order_by('title')

    if query:
        exams_queryset = exams_queryset.filter(
            Q(title__icontains=query) |
            Q(subject__name__icontains=query) |
            Q(subject__code__icontains=query)
        )
    results_list_early = Result.objects.filter(student=request.user).values_list('exam_id', flat=True)
    completed_exam_ids_early = set(results_list_early)

    show_only_pending = request.GET.get('pending') == '1'

    if show_only_pending:
        exams_queryset = exams_queryset.exclude(id__in=completed_exam_ids_early)
    subjects = Subject.objects.filter(
        exam__in=exams_queryset
    ).distinct().prefetch_related(
        Prefetch('exam_set', queryset=exams_queryset, to_attr='filtered_exams')
    ).order_by('name')

    results_list = Result.objects.filter(
        student=request.user
    ).select_related('exam', 'exam__subject').order_by('-submitted_at')

    completed_exam_ids = set(results_list.values_list('exam_id', flat=True))

    paginator = Paginator(results_list, 5)
    page_number = request.GET.get('page')
    results = paginator.get_page(page_number)

    profile = Profile.objects.filter(user=request.user).first()

    context = {
        'subjects': subjects,
        'results': results,
        'completed_exam_ids': completed_exam_ids,
        'query': query,
        'now': now,
        'profile': profile,
        'show_only_pending': show_only_pending,
    }

    return render(request, 'exams/dashboard.html', context)


DANGEROUS_MODULES = re.compile(
    r'\b(os|sys|subprocess|importlib|shutil|socket|__import__|builtins|eval|exec|open|compile)\b'
)

def run_python_code(code, input_data):
    
    if DANGEROUS_MODULES.search(code):
        return "Error: Use of restricted modules is not allowed."

    try:
        result = subprocess.run(
            [sys.executable, "-c", code],
            input=input_data,
            text=True,
            capture_output=True,
            timeout=3,
            env={"PATH": "/usr/bin:/bin"},
        )

        if result.stderr:
            return result.stderr.strip()

        return result.stdout.strip()

    except subprocess.TimeoutExpired:
        return "Time Limit Exceeded"


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def get_location(ip):
    try:
        response = requests.get(f"http://ip-api.com/json/{ip}", timeout=3)
        data = response.json()

        if data.get("status") == "success":
            return f"{data.get('country')}, {data.get('city')}"
        return "Unknown location"

    except Exception:
        return "Location fetch failed"


def normalize_sql_result(rows):
    normalized = []

    for row in rows:
        normalized_row = tuple(str(value).strip() for value in row)
        normalized.append(normalized_row)

    return sorted(normalized)


def run_sql_query(student_sql, setup_sql):

    stripped = student_sql.strip().rstrip(';')

    if not re.match(r'^\s*SELECT\b', stripped, re.IGNORECASE):
        return "SQL Error: Only SELECT statements are allowed."

    if ';' in stripped:
        return "SQL Error: Only a single statement is allowed."

    try:
        conn = sqlite3.connect(":memory:")
        cursor = conn.cursor()

        cursor.executescript(setup_sql)
        cursor.execute(stripped)

        rows = cursor.fetchall()
        conn.close()

        return normalize_sql_result(rows)

    except Exception as e:
        return f"SQL Error: {str(e)}"


@login_required
def take_exam(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)

    entered_code = request.POST.get("code") or request.session.get(f"exam_code_{exam_id}")
    session_key = f"exam_start_{exam_id}_{request.user.id}"

    if session_key not in request.session:
        request.session[session_key] = timezone.now().isoformat()

    exam_start_time = timezone.datetime.fromisoformat(request.session[session_key])
    if request.method == "GET":
        get_code = request.GET.get("code")
        if get_code:
            request.session[f"exam_code_{exam_id}"] = get_code
            entered_code = get_code

    if exam.access_code and entered_code != exam.access_code:
        messages.error(request, "Invalid exam access code.")
        return redirect('dashboard')

    profile = Profile.objects.filter(user=request.user).first()

    if not profile or profile.role != 'student':
        messages.error(request, "Only students can take exams.")
        return redirect('teacher_dashboard')

    if Result.objects.filter(student=request.user, exam=exam).exists():
        messages.warning(request, "You have already completed this exam.")
        return redirect('dashboard')

    now = timezone.now()

    if exam.start_time and now < exam.start_time:
        messages.warning(request, "This exam has not started yet.")
        return redirect('dashboard')

    if exam.end_time and now > exam.end_time:
        messages.error(request, "This exam deadline has passed.")
        return redirect('dashboard')

     
    questions = Question.objects.filter(exam=exam).order_by('id')

    if request.method == "POST":
        elapsed_minutes = (timezone.now() - exam_start_time).total_seconds() / 60
        grace_period = 2  

        if elapsed_minutes > exam.duration_minutes + grace_period:
            messages.error(request, "Time limit exceeded. Your exam could not be submitted.")
            request.session.pop(session_key, None)
            return redirect('dashboard')
        score = 0
        total = 0

        ip = get_client_ip(request)
        location = get_location(ip)

        if request.method == "POST":
            try:
                result = Result.objects.create(
                    student=request.user,
                    exam=exam,
                    score=0,
                    total=0,
                    ip_address=ip,
                    location=location
                )
            
            except IntegrityError:
                messages.warning(request, "You have already submitted this exam.")
                return redirect('dashboard')

        for question in questions:
            total += question.points

            selected_choice = None
            text_answer = ""
            code_answer = ""
            actual_output = ""
            passed_tests = 0
            total_tests = 0
            is_correct = False
            is_reviewed = False

            if question.question_type == "mcq":
                selected_choice_id = request.POST.get(f"question_{question.id}")

                if selected_choice_id:
                    selected_choice = Choice.objects.filter(id=selected_choice_id).first()

                    if selected_choice and selected_choice.is_correct:
                        score += question.points
                        is_correct = True

                is_reviewed = True

            elif question.question_type == "text":
                text_answer = request.POST.get(f"text_{question.id}", "")
                is_reviewed = False

            elif question.question_type == "code":
                code_answer = request.POST.get(f"code_{question.id}", "")

                test_cases = question.test_cases.all()
                total_tests = test_cases.count()
                output_logs = []

                for test in test_cases:
                    test_output = run_python_code(code_answer, test.input_data)

                    output_logs.append(
                        f"Input:\n{test.input_data}\n\n"
                        f"Expected Output:\n{test.expected_output.strip()}\n\n"
                        f"Actual Output:\n{test_output}\n"
                    )

                    if test_output.strip() == test.expected_output.strip():
                        passed_tests += 1

                actual_output = "\n\n--------------------\n\n".join(output_logs)

                if total_tests > 0:
                    earned_points = int((passed_tests / total_tests) * question.points)
                    score += earned_points

                is_correct = total_tests > 0 and passed_tests == total_tests
                is_reviewed = True

            elif question.question_type == "sql":
                code_answer = request.POST.get(f"code_{question.id}", "")

                test_cases = question.test_cases.all()
                total_tests = test_cases.count()
                output_logs = []

                for test in test_cases:
                    test_output = run_sql_query(code_answer, test.setup_sql)

                    try:
                        expected = ast.literal_eval(test.expected_output.strip())
                        expected = normalize_sql_result(expected)

                        if test_output == expected:
                            passed_tests += 1

                    except Exception:
                        expected = "Invalid expected output format"

                    output_logs.append(
                        f"Expected Output:\n{expected}\n\n"
                        f"Actual Output:\n{test_output}\n"
                    )

                actual_output = "\n\n--------------------\n\n".join(output_logs)

                if total_tests > 0:
                    earned_points = int((passed_tests / total_tests) * question.points)
                    score += earned_points

                is_correct = total_tests > 0 and passed_tests == total_tests
                is_reviewed = True

            StudentAnswer.objects.create(
                result=result,
                question=question,
                selected_choice=selected_choice,
                text_answer=text_answer,
                code_answer=code_answer,
                actual_output=actual_output,
                passed_tests=passed_tests,
                total_tests=total_tests,
                is_correct=is_correct,
                is_reviewed=is_reviewed
            )

        result.score = score
        result.total = total
        result.save()

        request.session.pop(f"exam_code_{exam_id}", None)
        request.session.pop(session_key, None)
        return redirect('result', result_id=result.id)

    elapsed_seconds = (timezone.now() - exam_start_time).total_seconds()
    remaining_seconds = max(0, (exam.duration_minutes * 60) - elapsed_seconds)
    return render(request, 'exams/take_exam.html', {
        'exam': exam,
        'questions': questions,
        'remaining_seconds': int(remaining_seconds)
    })


@login_required
def result_page(request, result_id):
    result = get_object_or_404(Result, id=result_id, student=request.user)
    answers = result.answers.all()

    return render(request, 'exams/result.html', {
        'result': result,
        'answers': answers
    })


@login_required
def teacher_dashboard(request):
    profile = Profile.objects.filter(user=request.user).first()

    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    query = request.GET.get('q', '').strip()

    total_students = Profile.objects.filter(role='student').count()
    total_exams = Exam.objects.count()
    total_results = Result.objects.count()

    average_score = Result.objects.aggregate(avg=Avg('score'))['avg']

    exam_stats = Exam.objects.annotate(
        attempts=Count('result'),
        avg_score=Avg('result__score')
    ).order_by('title')

    top_students = User.objects.filter(result__isnull=False).annotate(
        total_score=Sum('result__score'),
        exams_taken=Count('result')
    ).order_by('-total_score')[:5]

    recent_results = Result.objects.select_related(
        'student', 'exam', 'exam__subject'
    )

    if query:
        exam_stats = exam_stats.filter(
            Q(title__icontains=query) |
            Q(subject__name__icontains=query) |
            Q(subject__code__icontains=query)
        )

        recent_results = recent_results.filter(
            Q(student__username__icontains=query) |
            Q(exam__title__icontains=query) |
            Q(exam__subject__name__icontains=query) |
            Q(exam__subject__code__icontains=query)
        )

    
    sort_by = request.GET.get('sort', '-submitted_at')
    allowed_sorts = ['submitted_at', '-submitted_at', 'score', '-score', 'student__username', '-student__username']

    if sort_by not in allowed_sorts:
        sort_by = '-submitted_at'

    recent_results = recent_results.order_by(sort_by)

    results_paginator = Paginator(recent_results, 10)
    results_page_number = request.GET.get('results_page')
    recent_results = results_paginator.get_page(results_page_number)

    exam_labels_json = json.dumps([exam.title for exam in exam_stats])
    exam_scores_json = json.dumps([
        float(exam.avg_score or 0) for exam in exam_stats
    ])
    context = {
        'total_students': total_students,
        'total_exams': total_exams,
        'total_results': total_results,
        'average_score': round(average_score or 0, 2),
        'exam_stats': exam_stats,
        'top_students': top_students,
        'recent_results': recent_results,
        'query': query,
        'exam_labels_json': exam_labels_json,
        'exam_scores_json': exam_scores_json,
        'sort_by': sort_by,
    }

    return render(request, 'exams/teacher_dashboard.html', context)


@login_required
def export_results_csv(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="exam_results.csv"'

    writer = csv.writer(response)
    writer.writerow(['Student', 'Exam', 'Subject', 'Score', 'Total', 'Submitted At'])

    results = Result.objects.select_related('student', 'exam', 'exam__subject').all()

    for result in results:
        writer.writerow([
            result.student.username,
            result.exam.title,
            result.exam.subject.name,
            result.score,
            result.total,
            result.submitted_at
        ])

    return response


@login_required
def create_exam(request):
    profile = Profile.objects.filter(user=request.user).first()

    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    subjects = Subject.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        subject_id = request.POST.get('subject')
        duration = request.POST.get('duration', '').strip()
        start_time = request.POST.get('start_time') or None
        end_time = request.POST.get('end_time') or None
        access_code = request.POST.get('access_code', '').strip()

        errors = []

        if not title:
            errors.append("Exam title is required.")
        elif len(title) > 200:
            errors.append("Exam title is too long (max 200 characters).")

        if not subject_id:
            errors.append("You must select a subject.")

        subject = None
        if subject_id:
            subject = Subject.objects.filter(id=subject_id).first()
            if not subject:
                errors.append("Selected subject does not exist.")

        if not duration:
            errors.append("Duration is required.")
        else:
            try:
                duration = int(duration)
                if duration <= 0:
                    errors.append("Duration must be a positive number.")
                elif duration > 600:
                    errors.append("Duration cannot exceed 600 minutes.")
            except ValueError:
                errors.append("Duration must be a whole number.")
                duration = None

        if start_time and end_time:
            if end_time <= start_time:
                errors.append("End time must be after start time.")

        if access_code and len(access_code) > 30:
            errors.append("Access code is too long (max 30 characters).")

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'exams/create_exam.html', {
                'subjects': subjects,
                'form_data': request.POST,
            })

        Exam.objects.create(
            title=title,
            subject=subject,
            duration_minutes=duration,
            created_by=request.user,
            start_time=start_time,
            end_time=end_time,
            access_code=access_code,
            is_active=True
        )

        messages.success(request, f'Exam "{title}" created successfully.')
        return redirect('teacher_dashboard')

    return render(request, 'exams/create_exam.html', {'subjects': subjects})


@api_view(['GET'])
@drf_permission_classes([IsAuthenticated])
def subject_list(request):
    subjects = Subject.objects.all()
    serializer = SubjectSerializer(subjects, many=True)
    return Response(serializer.data)


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsTeacherOrReadOnly]  


class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [IsTeacherOrReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['subject', 'is_active']


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [IsTeacher]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['exam', 'question_type']


class ChoiceViewSet(viewsets.ModelViewSet):
    queryset = Choice.objects.all()
    serializer_class = ChoiceSerializer
    permission_classes = [IsTeacher]


class ResultViewSet(viewsets.ModelViewSet):
    serializer_class = ResultSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['exam']

    def get_queryset(self):
        if self.request.user.is_staff:
            return Result.objects.all()
        return Result.objects.filter(student=self.request.user)


class StudentAnswerViewSet(viewsets.ModelViewSet):
    queryset = StudentAnswer.objects.all()
    serializer_class = StudentAnswerSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.is_staff:
            return StudentAnswer.objects.all()
        profile = Profile.objects.filter(user=self.request.user).first()
        if profile and profile.role == 'teacher':
            return StudentAnswer.objects.filter(result__exam__created_by=self.request.user)
        return StudentAnswer.objects.filter(result__student=self.request.user)


class TestCaseViewSet(viewsets.ModelViewSet):
    queryset = TestCase.objects.all()
    serializer_class = TestCaseSerializer
    permission_classes = [IsTeacher]


class DashboardAPIView(APIView):
    permission_classes = [IsTeacher]

    def get(self, request):
        data = {
            "subjects": Subject.objects.count(),
            "exams": Exam.objects.count(),
            "questions": Question.objects.count(),
            "students": User.objects.count(),
            "results": Result.objects.count(),
        }

        return Response(data)


@login_required
def subject_list(request):
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != 'teacher':
        return redirect('dashboard')
    
    subjects = Subject.objects.all().order_by('name')
    return render(request, 'exams/subject_list.html', {'subjects': subjects})


@login_required
def subject_create(request):
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()

        if name and code:
            Subject.objects.create(name=name, code=code)
            messages.success(request, f'Subject "{name}" created successfully.')
            return redirect('subject_list')
        else:
            messages.error(request, 'Both name and code are required.')

    return render(request, 'exams/subject_form.html', {'action': 'Create'})


@login_required
def subject_edit(request, subject_id):
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    subject = get_object_or_404(Subject, id=subject_id)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()

        if name and code:
            subject.name = name
            subject.code = code
            subject.save()
            messages.success(request, f'Subject "{name}" updated successfully.')
            return redirect('subject_list')
        else:
            messages.error(request, 'Both name and code are required.')

    return render(request, 'exams/subject_form.html', {
        'action': 'Edit',
        'subject': subject
    })


@login_required
def subject_delete(request, subject_id):
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    subject = get_object_or_404(Subject, id=subject_id)

    if request.method == 'POST':
        name = subject.name
        subject.delete()
        messages.success(request, f'Subject "{name}" deleted.')
        return redirect('subject_list')

    return render(request, 'exams/subject_confirm_delete.html', {'subject': subject})


@login_required
def profile_view(request):
    profile = Profile.objects.filter(user=request.user).first()

    if not profile:
        profile = Profile.objects.create(user=request.user, role='student')

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        university_id = request.POST.get('university_id', '').strip()
        group_name = request.POST.get('group_name', '').strip()

        request.user.first_name = first_name
        request.user.last_name = last_name
        request.user.save()

        profile.university_id = university_id
        profile.group_name = group_name
        profile.save()

        messages.success(request, 'Profile updated successfully.')
        return redirect('profile')

    return render(request, 'exams/profile.html', {'profile': profile})


@login_required
def question_create(request, exam_id):
    profile = Profile.objects.filter(user=request.user).first()

    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    exam = get_object_or_404(Exam, id=exam_id)

    if exam.created_by != request.user:
        messages.error(request, "You don't have permission to manage this exam.")
        return redirect('teacher_dashboard')
    
    if request.method == 'POST':
        question_type = request.POST.get('question_type')
        text = request.POST.get('text', '').strip()
        points = request.POST.get('points', 1)

        if not text:
            messages.error(request, 'Question text is required.')
            return redirect('question_create', exam_id=exam.id)

        question = Question.objects.create(
            exam=exam,
            question_type=question_type,
            text=text,
            points=points
        )

        if question_type == 'mcq':
            choice_texts = request.POST.getlist('choice_text')
            correct_index = request.POST.get('correct_choice')

            for i, choice_text in enumerate(choice_texts):
                choice_text = choice_text.strip()
                if choice_text:
                    Choice.objects.create(
                        question=question,
                        text=choice_text,
                        is_correct=(str(i) == correct_index)
                    )

        messages.success(request, 'Question added successfully.')
        return redirect('question_create', exam_id=exam.id)

    questions = Question.objects.filter(exam=exam).order_by('id')

    return render(request, 'exams/question_create.html', {
        'exam': exam,
        'questions': questions
    })


@login_required
def grade_text_answers(request):
    profile = Profile.objects.filter(user=request.user).first()

    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    ungraded = StudentAnswer.objects.filter(
        question__question_type='text',
        is_reviewed=False,
        result__exam__created_by=request.user
    ).select_related('question', 'result', 'result__student', 'result__exam')

    return render(request, 'exams/grade_text_answers.html', {
        'ungraded': ungraded
    })


@login_required
def grade_single_answer(request, answer_id):
    profile = Profile.objects.filter(user=request.user).first()

    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    answer = get_object_or_404(StudentAnswer, id=answer_id)

    if answer.result.exam.created_by != request.user:
        messages.error(request, "You don't have permission to grade this answer.")
        return redirect('grade_text_answers')
    
    if request.method == 'POST':
        if answer.is_reviewed:
            messages.warning(request, 'This answer has already been graded.')
            return redirect('grade_text_answers')

        points_awarded = int(request.POST.get('points_awarded', 0))
        max_points = answer.question.points

        points_awarded = max(0, min(points_awarded, max_points))

        result = answer.result
        result.score += points_awarded
        result.save()

        answer.is_correct = (points_awarded == max_points)
        answer.is_reviewed = True
        answer.save()

        messages.success(request, 'Answer graded successfully.')
        return redirect('grade_text_answers')

    return render(request, 'exams/grade_single_answer.html', {'answer': answer})


@login_required
def question_import(request, exam_id):
    profile = Profile.objects.filter(user=request.user).first()

    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    exam = get_object_or_404(Exam, id=exam_id)

    if exam.created_by != request.user:
        messages.error(request, "You don't have permission to manage this exam.")
        return redirect('teacher_dashboard')
    
    if request.method == 'POST':
        uploaded_file = request.FILES.get('question_file')

        if not uploaded_file:
            messages.error(request, 'Please select a file to upload.')
            return redirect('question_import', exam_id=exam.id)

        filename = uploaded_file.name.lower()

        try:
            if filename.endswith('.csv'):
                rows = parse_csv_file(uploaded_file)
            elif filename.endswith('.xlsx'):
                rows = parse_xlsx_file(uploaded_file)
            else:
                messages.error(request, 'Only .csv and .xlsx files are supported.')
                return redirect('question_import', exam_id=exam.id)
        except Exception as e:
            messages.error(request, f'Could not read the file: {str(e)}')
            return redirect('question_import', exam_id=exam.id)

        created_count = 0
        error_rows = []

        for row_num, row in enumerate(rows, start=2):  
            try:
                question_type = row.get('question_type', '').strip().lower()
                text = row.get('text', '').strip()
                points = int(row.get('points') or 1)

                if not text or question_type not in ('mcq', 'text', 'code', 'sql'):
                    error_rows.append(row_num)
                    continue

                question = Question.objects.create(
                    exam=exam,
                    question_type=question_type,
                    text=text,
                    points=points
                )

                if question_type == 'mcq':
                    choices = [
                        row.get('choice_1', '').strip(),
                        row.get('choice_2', '').strip(),
                        row.get('choice_3', '').strip(),
                        row.get('choice_4', '').strip(),
                    ]
                    correct_index = row.get('correct_choice', '').strip()

                    has_valid_choice = False

                    for i, choice_text in enumerate(choices, start=1):
                        if choice_text:
                            is_correct = (str(i) == correct_index)
                            if is_correct:
                                has_valid_choice = True
                            Choice.objects.create(
                                question=question,
                                text=choice_text,
                                is_correct=is_correct
                            )

                    if not has_valid_choice:
                        error_rows.append(row_num)

                created_count += 1

            except Exception:
                error_rows.append(row_num)

        if created_count:
            messages.success(request, f'{created_count} question(s) imported successfully.')

        if error_rows:
            messages.warning(request, f'Rows with issues (skipped or incomplete): {", ".join(map(str, error_rows))}')

        return redirect('question_create', exam_id=exam.id)

    return render(request, 'exams/question_import.html', {'exam': exam})


def parse_csv_file(uploaded_file):
    decoded = uploaded_file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(decoded))
    return list(reader)


def parse_xlsx_file(uploaded_file):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() if h else '' for h in next(rows_iter)]

    rows = []
    for row_values in rows_iter:
        if all(v is None for v in row_values):
            continue
        row_dict = {}
        for header, value in zip(headers, row_values):
            row_dict[header] = str(value) if value is not None else ''
        rows.append(row_dict)

    return rows


@login_required
def question_analytics(request, exam_id):
    profile = Profile.objects.filter(user=request.user).first()

    if not profile or profile.role != 'teacher':
        return redirect('dashboard')

    exam = get_object_or_404(Exam, id=exam_id)

    if exam.created_by != request.user:
        messages.error(request, "You don't have permission to manage this exam.")
        return redirect('teacher_dashboard')
    
    questions = Question.objects.filter(exam=exam).order_by('id')

    question_stats = []

    for question in questions:
        answers = StudentAnswer.objects.filter(question=question)
        total_attempts = answers.count()

        if question.question_type in ('code', 'sql'):
            correct_count = answers.filter(is_correct=True).count()
        elif question.question_type == 'text':
            reviewed = answers.filter(is_reviewed=True)
            total_attempts = reviewed.count()
            correct_count = reviewed.filter(is_correct=True).count()
        else:
            correct_count = answers.filter(is_correct=True).count()

        if total_attempts > 0:
            pass_rate = round((correct_count / total_attempts) * 100, 1)
        else:
            pass_rate = None

        question_stats.append({
            'question': question,
            'total_attempts': total_attempts,
            'correct_count': correct_count,
            'pass_rate': pass_rate,
        })

    question_stats.sort(key=lambda x: (x['pass_rate'] is None, x['pass_rate'] if x['pass_rate'] is not None else 0))

    return render(request, 'exams/question_analytics.html', {
        'exam': exam,
        'question_stats': question_stats
    })


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsTeacherOrReadOnly]


class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [IsTeacherOrReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['subject', 'is_active']


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [IsTeacher]  
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['exam', 'question_type']


class ChoiceViewSet(viewsets.ModelViewSet):
    queryset = Choice.objects.all()
    serializer_class = ChoiceSerializer
    permission_classes = [IsTeacher]  


class TestCaseViewSet(viewsets.ModelViewSet):
    queryset = TestCase.objects.all()
    serializer_class = TestCaseSerializer
    permission_classes = [IsTeacher]  


class DashboardAPIView(APIView):
    permission_classes = [IsTeacher]  